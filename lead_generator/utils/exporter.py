"""
Data Exporter v3
================
Exports lead data to CSV and Excel with professional design.
Full dashboard with charts, alternating rows, column groups, breakdowns.
4-sheet Excel: Dashboard, Leads, By Country, By Category.
"""

import os
import csv
from typing import List, Optional, Dict, Any
from datetime import datetime
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

from lead_generator.models.lead import Lead, LeadCollection
from lead_generator.config import ScraperConfig


COLORS = {
    "header_bg": "1B2A4A", "header_bg2": "2D4373",
    "group_bg": "D6E4F0", "row_even": "FFFFFF", "row_odd": "F4F6F9",
    "hot": "E2EFDA", "warm": "FFF2CC", "cold": "FCE4EC",
    "hot_text": "2E7D32", "warm_text": "E65100", "cold_text": "C62828",
    "title_bg": "0D1B2A", "subtitle_bg": "1B3A5C", "credit_gray": "888888",
    "chart_green": "4CAF50", "chart_yellow": "FFC107", "chart_red": "EF5350",
    "chart_blue": "42A5F5", "chart_orange": "FF7043", "chart_light": "BDBDBD",
    "chart_teal": "26A69A", "chart_purple": "AB47BC",
    "accent": "1E88E5", "light_gray": "F5F5F5",
}

COLUMN_GROUPS = [
    ("Business Info", ["Business Name", "Category", "Industry", "Rating",
                       "Review Count", "Price Level", "Operating Hours", "Description"]),
    ("Contact", ["Phone", "Email", "All Emails", "Website"]),
    ("Location", ["Address", "City", "State", "Zip Code", "Country"]),
    ("Social Media", ["Facebook", "Instagram", "LinkedIn", "Twitter", "YouTube",
                      "TikTok", "WhatsApp", "Telegram", "Snapchat", "Pinterest", "TripAdvisor"]),
    ("Decision Makers", ["Owner Name", "Manager Name"]),
    ("Meta", ["Source", "Google Maps URL", "Scraped At", "Lead Score", "Data Completeness"]),
]

HEADERS = [
    "Business Name", "Phone", "Email", "All Emails", "Website",
    "Address", "City", "State", "Zip Code", "Country",
    "Category", "Industry", "Rating", "Review Count", "Price Level",
    "Operating Hours", "Description",
    "Facebook", "Instagram", "LinkedIn", "Twitter", "YouTube", "TikTok",
    "WhatsApp", "Telegram", "Snapchat", "Pinterest", "TripAdvisor",
    "Owner Name", "Manager Name",
    "Source", "Google Maps URL", "Scraped At", "Lead Score", "Data Completeness",
]

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class LeadExporter:

    def __init__(self, config=None):
        self.config = config or ScraperConfig()
        self.output_dir = self.config.output_directory
        os.makedirs(self.output_dir, exist_ok=True)

    def export(self, leads=None, filename=None, fmt=None):
        if not leads:
            print("[-] No leads to export")
            return []
        fmt = fmt or self.config.output_format
        filename = filename or self._generate_filename()
        files = []
        if fmt in ("csv", "both"):
            files.append(self._export_csv(leads, filename))
        if fmt in ("excel", "both"):
            files.append(self._export_excel(leads, filename))
        self._print_summary(leads)
        return files

    def _export_csv(self, leads, filename):
        filepath = os.path.join(self.output_dir, f"{filename}.csv")
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=HEADERS)
            writer.writerow({h: h.upper() for h in HEADERS})
            for lead in leads:
                writer.writerow(lead.to_csv_row())
        print(f"[+] CSV exported: {filepath} ({len(leads)} leads)")
        return filepath

    def _export_excel(self, leads, filename):
        filepath = os.path.join(self.output_dir, f"{filename}.xlsx")
        wb = Workbook()

        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        self._write_dashboard(ws_dash, leads, wb)

        ws_leads = wb.create_sheet("Leads")
        self._write_leads_sheet(ws_leads, leads)

        ws_country = wb.create_sheet("By Country")
        self._write_breakdown_sheet(ws_country, leads, "country")

        ws_category = wb.create_sheet("By Category")
        self._write_breakdown_sheet(ws_category, leads, "category")

        credit = "Made by Mustapha Elasri | github.com/Stoph1723"
        for ws in wb.worksheets:
            ws.oddFooter.center.text = credit
            ws.evenFooter.center.text = credit
            ws.sheet_properties.tabColor = "1E88E5"

        wb.save(filepath)
        print(f"[+] Excel exported: {filepath} ({len(leads)} leads)")
        return filepath

    def _write_dashboard(self, ws, leads, wb):
        stats = leads.stats
        total = stats["total"]
        if total == 0:
            ws.cell(row=1, column=1, value="No leads to display")
            return

        ws.merge_cells("A1:L1")
        c = ws["A1"]
        c.value = "  LEAD GENERATION DASHBOARD"
        c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        c.fill = PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:L2")
        c = ws["A2"]
        c.value = f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Total Leads: {total}"
        c.font = Font(color="FFFFFF", size=9)
        c.fill = PatternFill(start_color=COLORS["subtitle_bg"], end_color=COLORS["subtitle_bg"], fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        self._make_metric_cards(ws, stats, total)

        ws_data = wb.create_sheet("_ChartData")
        ws_data.sheet_state = "hidden"

        tier_labels = ["Hot (70+)", "Warm (40-69)", "Cold (<40)"]
        tier_values = [stats["hot_leads"], stats["warm_leads"], stats["cold_leads"]]
        tier_colors = [COLORS["chart_green"], COLORS["chart_yellow"], COLORS["chart_red"]]
        ws_data.cell(row=1, column=1, value="Tier")
        ws_data.cell(row=1, column=2, value="Count")
        for i, (lbl, val) in enumerate(zip(tier_labels, tier_values)):
            ws_data.cell(row=2 + i, column=1, value=lbl)
            ws_data.cell(row=2 + i, column=2, value=val)

        with_both = sum(1 for l in leads if l.email and l.phone)
        with_email_only = sum(1 for l in leads if l.email and not l.phone)
        with_phone_only = sum(1 for l in leads if l.phone and not l.email)
        with_neither = total - with_both - with_email_only - with_phone_only
        contact_labels = ["Email + Phone", "Email Only", "Phone Only", "Neither"]
        contact_values = [with_both, with_email_only, with_phone_only, with_neither]
        contact_colors = [COLORS["chart_green"], COLORS["chart_blue"], COLORS["chart_orange"], COLORS["chart_light"]]
        ws_data.cell(row=6, column=1, value="Contact")
        ws_data.cell(row=6, column=2, value="Count")
        for i, (lbl, val) in enumerate(zip(contact_labels, contact_values)):
            ws_data.cell(row=7 + i, column=1, value=lbl)
            ws_data.cell(row=7 + i, column=2, value=val)

        source_counts = Counter(l.source for l in leads)
        source_labels = [s for s, _ in source_counts.most_common(10)]
        source_values = [c for _, c in source_counts.most_common(10)]
        ws_data.cell(row=12, column=1, value="Source")
        ws_data.cell(row=12, column=2, value="Count")
        for i, (lbl, val) in enumerate(zip(source_labels, source_values)):
            ws_data.cell(row=13 + i, column=1, value=lbl)
            ws_data.cell(row=13 + i, column=2, value=val)
        source_end = 12 + len(source_labels)

        country_counts = Counter(l.country for l in leads if l.country)
        if not country_counts:
            country_counts = Counter("Unknown" for _ in leads)
        country_labels = [c for c, _ in country_counts.most_common(10)]
        country_values = [v for _, v in country_counts.most_common(10)]
        ws_data.cell(row=25, column=1, value="Country")
        ws_data.cell(row=25, column=2, value="Count")
        for i, (lbl, val) in enumerate(zip(country_labels, country_values)):
            ws_data.cell(row=26 + i, column=1, value=lbl)
            ws_data.cell(row=26 + i, column=2, value=val)
        country_end = 25 + len(country_labels)

        chart_w = 9
        chart_h = 7

        pie1 = PieChart()
        pie1.title = "Lead Quality"
        pie1.style = 10
        pie1.width = chart_w
        pie1.height = chart_h
        pie1.add_data(Reference(ws_data, min_col=2, min_row=1, max_row=4), titles_from_data=True)
        pie1.set_categories(Reference(ws_data, min_col=1, min_row=2, max_row=4))
        pie1.dataLabels = DataLabelList()
        pie1.dataLabels.showPercent = True
        pie1.dataLabels.showVal = False
        pie1.dataLabels.showCatName = False
        for idx in range(len(tier_labels)):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = tier_colors[idx]
            pie1.series[0].data_points.append(pt)
        ws.add_chart(pie1, "A6")

        pie2 = PieChart()
        pie2.title = "Contact Coverage"
        pie2.style = 10
        pie2.width = chart_w
        pie2.height = chart_h
        pie2.add_data(Reference(ws_data, min_col=2, min_row=6, max_row=10), titles_from_data=True)
        pie2.set_categories(Reference(ws_data, min_col=1, min_row=7, max_row=10))
        pie2.dataLabels = DataLabelList()
        pie2.dataLabels.showPercent = True
        pie2.dataLabels.showVal = False
        pie2.dataLabels.showCatName = False
        for idx in range(len(contact_labels)):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = contact_colors[idx]
            pie2.series[0].data_points.append(pt)
        ws.add_chart(pie2, "G6")

        bar1 = BarChart()
        bar1.type = "col"
        bar1.title = "Leads by Source"
        bar1.style = 10
        bar1.width = chart_w
        bar1.height = chart_h
        bar1.y_axis.title = "Count"
        bar1.add_data(Reference(ws_data, min_col=2, min_row=12, max_row=source_end), titles_from_data=True)
        bar1.set_categories(Reference(ws_data, min_col=1, min_row=13, max_row=source_end))
        bar1.shape = 4
        if bar1.series:
            bar1.series[0].graphicalProperties.solidFill = COLORS["chart_blue"]
        ws.add_chart(bar1, "A14")

        bar2 = BarChart()
        bar2.type = "col"
        bar2.title = "Leads by Country"
        bar2.style = 10
        bar2.width = chart_w
        bar2.height = chart_h
        bar2.y_axis.title = "Count"
        bar2.add_data(Reference(ws_data, min_col=2, min_row=25, max_row=country_end), titles_from_data=True)
        bar2.set_categories(Reference(ws_data, min_col=1, min_row=26, max_row=country_end))
        bar2.shape = 4
        if bar2.series:
            bar2.series[0].graphicalProperties.solidFill = COLORS["chart_orange"]
        ws.add_chart(bar2, "G14")

        credit_font = Font(italic=True, color=COLORS["credit_gray"], size=8)
        ws.cell(row=22, column=1, value="Made by Mustapha Elasri | github.com/Stoph1723")
        ws.cell(row=22, column=1).font = credit_font
        ws.cell(row=22, column=1).protection = Protection(locked=True)

    def _make_metric_cards(self, ws, stats, total):
        accent = PatternFill(start_color=COLORS["accent"], end_color=COLORS["accent"], fill_type="solid")
        green = PatternFill(start_color=COLORS["chart_green"], end_color=COLORS["chart_green"], fill_type="solid")
        orange = PatternFill(start_color=COLORS["chart_orange"], end_color=COLORS["chart_orange"], fill_type="solid")
        red = PatternFill(start_color=COLORS["chart_red"], end_color=COLORS["chart_red"], fill_type="solid")
        teal = PatternFill(start_color=COLORS["chart_teal"], end_color=COLORS["chart_teal"], fill_type="solid")

        card_f = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        card_h = 24

        metrics = [
            ("A", "B", "Total", total, accent),
            ("C", "D", "Avg Score", f'{stats["avg_score"]}', teal),
            ("E", "F", "Email", stats["with_email"], green),
            ("G", "H", "Phone", stats["with_phone"], orange),
            ("I", "J", "Hot", stats["hot_leads"], green),
            ("K", "L", "Warm", stats["warm_leads"], orange),
        ]
        for c1, c2, label, value, color in metrics:
            ws.merge_cells(f"{c1}4:{c2}4")
            cell = ws[f"{c1}4"]
            cell.value = f"{label}: {value}"
            cell.font = card_f
            cell.fill = color
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            ws.cell(row=4, column=ord(c2) - ord("A") + 1).border = thin_border
        ws.row_dimensions[4].height = card_h

    def _write_leads_sheet(self, ws, leads):
        thick_border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium"),
        )

        gf = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        gfill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        col_idx = 1
        for gname, gheaders in COLUMN_GROUPS:
            sc = col_idx
            ec = col_idx + len(gheaders) - 1
            if sc == ec:
                cell = ws.cell(row=1, column=sc, value=gname.upper())
            else:
                ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
                cell = ws.cell(row=1, column=sc, value=gname.upper())
            cell.font = gf
            cell.fill = gfill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
            for c in range(sc, ec + 1):
                ws.cell(row=1, column=c).fill = gfill
                ws.cell(row=1, column=c).border = thick_border
            col_idx = ec + 1

        hf = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        hfill = PatternFill(start_color=COLORS["accent"], end_color=COLORS["accent"], fill_type="solid")
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=2, column=ci, value=h.upper())
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thick_border
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 32

        ef = PatternFill(start_color=COLORS["row_even"], end_color=COLORS["row_even"], fill_type="solid")
        of = PatternFill(start_color=COLORS["row_odd"], end_color=COLORS["row_odd"], fill_type="solid")
        hot_f = PatternFill(start_color=COLORS["hot"], end_color=COLORS["hot"], fill_type="solid")
        warm_f = PatternFill(start_color=COLORS["warm"], end_color=COLORS["warm"], fill_type="solid")
        cold_f = PatternFill(start_color=COLORS["cold"], end_color=COLORS["cold"], fill_type="solid")

        for ri, lead in enumerate(leads):
            er = ri + 3
            rd = lead.to_csv_row()
            base = ef if ri % 2 == 0 else of

            for ci, h in enumerate(HEADERS, 1):
                cell = ws.cell(row=er, column=ci, value=rd.get(h, ""))
                cell.border = thin_border
                cell.fill = base
                cell.alignment = Alignment(vertical="center", wrap_text=(h in ("Description", "Operating Hours")))

                if h == "Lead Score":
                    try:
                        score = int(rd.get("Lead Score", "0"))
                        if score >= 70:
                            cell.fill = hot_f
                            cell.font = Font(bold=True, color=COLORS["hot_text"])
                        elif score >= 40:
                            cell.fill = warm_f
                            cell.font = Font(bold=True, color=COLORS["warm_text"])
                        else:
                            cell.fill = cold_f
                            cell.font = Font(bold=True, color=COLORS["cold_text"])
                    except ValueError:
                        pass

                if h in ("Email", "Website", "Google Maps URL", "Facebook", "Instagram",
                         "LinkedIn", "Twitter", "YouTube", "TikTok", "WhatsApp",
                         "Telegram", "Snapchat", "Pinterest", "TripAdvisor"):
                    val = rd.get(h, "")
                    if val and val.startswith("http"):
                        cell.hyperlink = val
                        cell.font = Font(color="0563C1", underline="single")

        for ci, h in enumerate(HEADERS, 1):
            ml = len(h)
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=ci, max_col=ci):
                for cell in row:
                    if cell.value:
                        ml = max(ml, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 2, 40)

        ws.freeze_panes = "A3"
        cr = ws.max_row + 2
        credit_font = Font(italic=True, color=COLORS["credit_gray"], size=8)
        ws.cell(row=cr, column=1, value="Made by Mustapha Elasri | github.com/Stoph1723")
        ws.cell(row=cr, column=1).font = credit_font
        ws.cell(row=cr, column=1).protection = Protection(locked=True)

    def _write_breakdown_sheet(self, ws, leads, field):
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = f"  BREAKDOWN BY {field.upper()}"
        t.font = Font(bold=True, size=14, color="FFFFFF")
        t.fill = PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        bds_h = [field.title(), "Total Leads", "Avg Score", "With Email", "With Phone", "Hot Leads"]
        hf = Font(bold=True, color="FFFFFF", size=10)
        hfill = PatternFill(start_color=COLORS["header_bg2"], end_color=COLORS["header_bg2"], fill_type="solid")
        for ci, h in enumerate(bds_h, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        groups = {}
        for lead in leads:
            key = getattr(lead, field, "") or "Unknown"
            groups.setdefault(key, []).append(lead)
        sg = sorted(groups.items(), key=lambda x: len(x[1]), reverse=True)

        ef = PatternFill(start_color=COLORS["row_even"], end_color=COLORS["row_even"], fill_type="solid")
        of_ = PatternFill(start_color=COLORS["row_odd"], end_color=COLORS["row_odd"], fill_type="solid")

        for ri, (gn, gl) in enumerate(sg):
            row = ri + 3
            fill = ef if ri % 2 == 0 else of_
            avg = round(sum(l.lead_score for l in gl) / len(gl), 1)
            we = sum(1 for l in gl if l.email)
            wp = sum(1 for l in gl if l.phone)
            ht = sum(1 for l in gl if l.lead_tier == "HOT")
            for ci, val in enumerate([gn, len(gl), avg, we, wp, ht], 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.border = thin_border
                cell.fill = fill
                if ci >= 3:
                    cell.alignment = Alignment(horizontal="center")

        for ci in range(1, len(bds_h) + 1):
            ml = len(bds_h[ci - 1])
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=ci, max_col=ci):
                for cell in row:
                    if cell.value:
                        ml = max(ml, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 3, 30)

        ws.freeze_panes = "A3"
        cr = ws.max_row + 2
        credit_font = Font(italic=True, color=COLORS["credit_gray"], size=8)
        ws.cell(row=cr, column=1, value="Made by Mustapha Elasri | github.com/Stoph1723")
        ws.cell(row=cr, column=1).font = credit_font
        ws.cell(row=cr, column=1).protection = Protection(locked=True)

    def _generate_filename(self):
        return f"{self.config.output_filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    def _print_summary(self, leads):
        stats = leads.stats
        print("\n" + "=" * 50)
        print("EXPORT SUMMARY")
        print("=" * 50)
        print(f"  Total Leads:    {stats['total']}")
        print(f"  With Email:     {stats['with_email']}")
        print(f"  With Phone:     {stats['with_phone']}")
        print(f"  Hot Leads:      {stats['hot_leads']}")
        print(f"  Warm Leads:     {stats['warm_leads']}")
        print(f"  Cold Leads:     {stats['cold_leads']}")
        print(f"  Average Score:  {stats['avg_score']}")
        print("=" * 50)
